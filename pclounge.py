import sys
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QPushButton,
    QVBoxLayout,
    QHBoxLayout,
    QWidget,
    QLabel,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QDialog,
    QLineEdit,
    QFormLayout,
    QMessageBox,
    QDialogButtonBox,
    QComboBox,
    QListWidget,
    QGridLayout,
    QCompleter,
)
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import QDateTime, Qt
import openpyxl
from openpyxl import Workbook
import resources_rc
import os

pcnums = [16,15,14,11,12,13,10,9,8,5,6,7,4,3,2,1]

class PCLoungeApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PC Lounge")

        self.setWindowIcon(QIcon(":/icons/myico.ico"))

        self.setGeometry(100, 100, 1220, 600)

        # Global setting for allowing same person to sign into the same PC
        self.allow_same_person = False

        # Create tab widget
        self.tab_widget = QTabWidget()
        self.main_tab = QWidget()
        self.club_people_tab = QWidget()
        self.tab_widget.addTab(self.main_tab, "Main")
        self.tab_widget.addTab(self.club_people_tab, "Club People")

        # Set up Main Tab
        self.setup_main_tab()

        # Set up Club People Tab
        self.setup_club_people_tab()

        # Add tab widget to the main layout
        main_layout = QHBoxLayout()
        main_layout.addWidget(self.tab_widget)
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        # Initialize Excel files
        self.log_file = self.init_log_excel_file()
        self.club_file = self.init_club_excel_file()

        # Load club members and events
        self.load_club_members()
        self.load_events()

    def setup_main_tab(self):
        layout = QHBoxLayout()

        # Left side: PC Grid
        self.pc_layout = QGridLayout()
        self.pc_statuses = (
            {}
        )  # Track whether each PC is free or in use, and store the row in Excel for each PC
        self.create_pc_grid(16)
        layout.addLayout(self.pc_layout)

        # Right side: Events
        right_layout = QVBoxLayout()
        self.people_table = QTableWidget()
        self.people_table.setColumnCount(4)
        self.people_table.setHorizontalHeaderLabels(
            ["PC Number", "Name", "Action", "Time"]
        )
        self.people_table.setColumnWidth(3, 200)
        self.people_table.verticalHeader().setVisible(False)

        right_layout.addWidget(QLabel("Events"))

        right_layout.addWidget(self.people_table)

        layout.addLayout(right_layout)

        self.main_tab.setLayout(layout)

    def setup_club_people_tab(self):
        layout = QVBoxLayout()

        # Add button
        self.add_button = QPushButton("Add Person")
        self.add_button.clicked.connect(self.add_person)
        layout.addWidget(self.add_button)

        # Delete button
        self.delete_button = QPushButton("Delete Person")
        self.delete_button.clicked.connect(self.delete_person)
        layout.addWidget(self.delete_button)

        # List of members
        self.club_people_list = QListWidget()
        layout.addWidget(QLabel("Club Members"))
        layout.addWidget(self.club_people_list)

        self.club_people_tab.setLayout(layout)

    def create_pc_grid(self, num_pcs):
        for i in range(num_pcs):
            pc_button = QPushButton(f"PC {pcnums[i]} (Free)")
            pc_button.setStyleSheet(
                "background-color: lightgreen; font-size: 16px; padding: 10px;"
            )
            pc_button.clicked.connect(lambda _, pc=i: self.toggle_pc_status(pc))

            pc_label = QLabel()

            if i < 12:  # First 3 rows of 4
                row, col = divmod(i, 3)
            else:  # Last row of 4
                row, col = 4, i - 12  # Set to the last row
            row = row*2

            self.pc_layout.addWidget(pc_button, row, col * 2)
            self.pc_layout.addWidget(pc_label, row+1, col * 2)

            self.pc_statuses[i] = {
                "status": "Free",
                "excel_row": None,
                "user_label": pc_label,
                "user_name":None,
            }

    def toggle_pc_status(self, pc_id):
        if self.pc_statuses[pc_id]["status"] == "Free":
            if not self.club_people_list.count():
                QMessageBox.warning(
                    self,
                    "No Members",
                    "No club members available. Please add members first.",
                )
                return

            # Prompt for club member selection
            dialog = QDialog(self)
            dialog.setWindowTitle("Select Club Member")

            layout = QVBoxLayout()

            self.member_combobox = QComboBox()
            self.member_combobox.setMinimumWidth(400)
            self.member_combobox.setEditable(True)  # Enable text input for searching

            # Populate the combo box with member names
            member_names = [self.club_people_list.item(i).text() for i in range(self.club_people_list.count())]
            self.member_combobox.addItems(member_names)

            # Enable filtering while typing
            completer = QCompleter(member_names)
            completer.setCaseSensitivity(Qt.CaseInsensitive)  # Case-insensitive search
            completer.setFilterMode(Qt.MatchContains)  # Match anywhere in the text
            self.member_combobox.setCompleter(completer)

            layout.addWidget(QLabel("Select Club Member:"))
            layout.addWidget(self.member_combobox)

            buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)

            # Validate selection before accepting
            def validate_selection():
                selected_member = self.member_combobox.currentText()
                if selected_member not in member_names:
                    QMessageBox.warning(dialog, "Invalid Selection", "Please select a valid club member.")
                else:
                    self.sign_in_member(pc_id, dialog)  # Proceed if valid

            buttons.accepted.connect(validate_selection)
            buttons.rejected.connect(dialog.reject)
            layout.addWidget(buttons)

            dialog.setLayout(layout)
            dialog.exec_()

        elif self.pc_statuses[pc_id]["status"] == "In Use":
            sign_out_time = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss")

            pc_button = self.pc_layout.itemAt(pc_id * 2).widget()
            pc_button.setText(f"PC {pcnums[pc_id]} (Free)")
            pc_button.setStyleSheet(
                "background-color: lightgreen; font-size: 16px; padding: 10px;"
            )
            self.pc_statuses[pc_id]["status"] = "Free"
            previous_user = self.pc_statuses[pc_id]["user_name"]
            self.pc_statuses[pc_id]["user_label"].setText("None")

            # Log sign-out
            self.log_sign_out_to_excel(
                self.pc_statuses[pc_id]["excel_row"], sign_out_time
            )

            # Update table for sign-out
            self.people_table.insertRow(self.people_table.rowCount())
            self.people_table.setItem(
                self.people_table.rowCount() - 1, 0, QTableWidgetItem(str(pcnums[pc_id]))
            )
            self.people_table.setItem(
                self.people_table.rowCount() - 1,
                1,
                QTableWidgetItem(str(previous_user)),
            )
            self.people_table.setItem(
                self.people_table.rowCount() - 1, 2, QTableWidgetItem("Sign-Out")
            )
            self.people_table.setItem(
                self.people_table.rowCount() - 1, 3, QTableWidgetItem(sign_out_time)
            )

    def sign_in_member(self, pc_id, dialog):
        selected_member = self.member_combobox.currentText()
        student_id = selected_member.split("(")[-1].strip(")")
        student_name = selected_member.split("(")[0].strip()
        sign_in_time = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss")

        # Check if the person is already signed in to the same PC
        if not self.allow_same_person:
            for pc, status in self.pc_statuses.items():
                if (
                    status["status"] == "In Use"
                    and f"Name: {student_name}\nID: {student_id}"
                    == status.get("user_label", "").text()
                ):
                    QMessageBox.warning(
                        self,
                        "Error",
                        f"{student_name} is already signed into another PC.",
                    )
                    dialog.reject()
                    return

        pc_button = self.pc_layout.itemAt(pc_id * 2).widget()
        pc_button.setText(f"PC {pcnums[pc_id]} (In Use)")
        pc_button.setStyleSheet(
            "background-color: lightcoral; font-size: 16px; padding: 10px;"
        )
        self.pc_statuses[pc_id]["status"] = "In Use"
        self.pc_statuses[pc_id]["user_label"].setText(f"Name: {student_name}\nID: {student_id}")
        self.pc_statuses[pc_id]["user_name"] = student_name

        self.people_table.insertRow(self.people_table.rowCount())
        self.people_table.setItem(
            self.people_table.rowCount() - 1, 0, QTableWidgetItem(str(pcnums[pc_id]))
        )
        self.people_table.setItem(
            self.people_table.rowCount() - 1, 1, QTableWidgetItem(student_name)
        )
        self.people_table.setItem(
            self.people_table.rowCount() - 1, 2, QTableWidgetItem("Sign-In")
        )
        self.people_table.setItem(
            self.people_table.rowCount() - 1, 3, QTableWidgetItem(sign_in_time)
        )

        row = self.log_sign_in_to_excel(
            student_name, student_id, sign_in_time, pcnums[pc_id]
        )
        self.pc_statuses[pc_id]["excel_row"] = row

        dialog.accept()

    def init_log_excel_file(self):
        current_time = QDateTime.currentDateTime().toString("yyyyMMdd")
        log_file = f"PC_Lounge_Log_{current_time}.xlsx"

        if not os.path.exists(log_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "PC Log"
            ws.append(
                ["Name", "Student ID", "Sign-In Time", "PC Number", "Sign-Out Time"]
            )
            wb.save(log_file)

        return log_file

    def log_sign_in_to_excel(self, name, student_id, sign_in_time, pc_number):
        wb = openpyxl.load_workbook(self.log_file)
        ws = wb.active
        ws.append([name, student_id, sign_in_time, pc_number, ""])
        row = ws.max_row
        wb.save(self.log_file)
        wb.close()
        return row

    def log_sign_out_to_excel(self, row, sign_out_time):
        wb = openpyxl.load_workbook(self.log_file)
        ws = wb.active
        ws.cell(row=row, column=5).value = sign_out_time
        wb.save(self.log_file)
        wb.close()

    def init_club_excel_file(self):
        club_file = "Club_Members.xlsx"

        if not os.path.exists(club_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Members"
            ws.append(["Student ID", "Name"])
            wb.save(club_file)

        return club_file

    def load_club_members(self):
        self.club_people_list.clear()
        wb = openpyxl.load_workbook(self.club_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.club_people_list.addItem(f"{row[1]} ({row[0]})")
        wb.close()

    def load_events(self):
        current_date = QDateTime.currentDateTime().toString("yyyy-MM-dd")
        self.people_table.setRowCount(0)
        wb = openpyxl.load_workbook(self.log_file)
        ws = wb.active

        events = []  # Store events as tuples for sorting

        for row in ws.iter_rows(min_row=2, values_only=True):
            sign_in_time = row[2]
            sign_out_time = row[4]
            if sign_in_time and sign_in_time.startswith(current_date):
                events.append((str(row[3]), row[0], "Sign-In", sign_in_time))
            if sign_out_time and sign_out_time.startswith(current_date):
                events.append((str(row[3]), row[0], "Sign-Out", sign_out_time))

        wb.close()

        # Sort events by the timestamp (4th item in tuple)
        events.sort(key=lambda x: QDateTime.fromString(x[3], "yyyy-MM-dd hh:mm:ss"))

        # Insert sorted events into the table
        for event in events:
            row_position = self.people_table.rowCount()
            self.people_table.insertRow(row_position)
            for col, value in enumerate(event):
                self.people_table.setItem(row_position, col, QTableWidgetItem(value))

        # Ensure sorting applies correctly within the table widget
        self.people_table.sortItems(3, Qt.AscendingOrder)

    def add_person(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Person")

        layout = QFormLayout()
        self.student_id_input = QLineEdit()
        self.name_input = QLineEdit()
        layout.addRow("Student ID:", self.student_id_input)
        layout.addRow("Name:", self.name_input)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.save_person)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        dialog.setLayout(layout)
        dialog.exec_()

    def save_person(self):
        student_id = self.student_id_input.text()
        name = self.name_input.text()

        if student_id and name:
            wb = openpyxl.load_workbook(self.club_file)
            ws = wb.active
            ws.append([student_id, name])
            wb.save(self.club_file)
            wb.close()
            self.club_people_list.addItem(f"{name} ({student_id})")
        else:
            QMessageBox.warning(
                self, "Input Error", "Both Student ID and Name are required."
            )

    def delete_person(self):
        selected_item = self.club_people_list.currentItem()
        if selected_item:
            student_id = selected_item.text().split("(")[-1].strip(")")
            self.remove_person_from_file(student_id)
            self.club_people_list.takeItem(self.club_people_list.row(selected_item))

    def remove_person_from_file(self, student_id):
        wb = openpyxl.load_workbook(self.club_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        ws.delete_rows(2, ws.max_row - 1)  # Clear all rows except the header

        for row in rows[1:]:
            if row[0] != student_id:
                ws.append(row)

        wb.save(self.club_file)
        wb.close()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyleSheet(
        """
    QMainWindow {
        background-color: black;
        color: white;
    }
    QPushButton {
        color: black;
        border: 1px solid #5a5a5a;
        border-radius: 5px;
        padding: 8px 16px;
    }
    QPushButton:hover {
        background-color: #505050;
    }
    QLabel {
        font-size: 14px;
        color: black;
    }
    QTableWidget {
        background-color: black;
        color: white;
        gridline-color: #5a5a5a;
    }
"""
    )

    window = PCLoungeApp()
    window.show()
    sys.exit(app.exec_())
