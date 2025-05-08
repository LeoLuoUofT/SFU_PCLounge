import sys
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QPushButton,
    QVBoxLayout,
    QHBoxLayout,
    QWidget,
    QLabel,
    QListWidget,
    QTabWidget,
    QDialog,
    QLineEdit,
    QFormLayout,
    QMessageBox,
    QDialogButtonBox,
    QComboBox,
    QVBoxLayout,
)
from PyQt5.QtCore import QDateTime
import openpyxl
from openpyxl import Workbook
import os


class PCLoungeApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PC Lounge")
        self.setGeometry(100, 100, 800, 600)

        # Create tab widget
        self.tab_widget = QTabWidget()
        self.main_tab = QWidget()
        self.club_people_tab = QWidget()
        self.tab_widget.addTab(self.main_tab, "Main")
        self.tab_widget.addTab(self.club_people_tab, "Club People")

        # Set up Main Tab
        self.setup_main_tab()

        # Set up Club People Tab
        self.setup_club_people_tab()

        # Add tab widget to the main layout
        main_layout = QHBoxLayout()
        main_layout.addWidget(self.tab_widget)
        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

        # Initialize Excel files
        self.log_file = self.init_log_excel_file()
        self.club_file = self.init_club_excel_file()

        # Load club members
        self.load_club_members()

    def setup_main_tab(self):
        layout = QHBoxLayout()

        # Left side: PC Grid
        self.pc_layout = QVBoxLayout()
        self.pc_statuses = (
            {}
        )  # Track whether each PC is free or in use, and store the row in Excel for each PC
        self.create_pc_grid(10)  # 10 PCs for example
        layout.addLayout(self.pc_layout)

        # Right side: Registered people
        right_layout = QVBoxLayout()
        self.people_list = QListWidget()
        right_layout.addWidget(QLabel("Registered People"))
        right_layout.addWidget(self.people_list)

        layout.addLayout(right_layout)

        self.main_tab.setLayout(layout)

    def setup_club_people_tab(self):
        layout = QVBoxLayout()

        # Add button
        self.add_button = QPushButton("Add Person")
        self.add_button.clicked.connect(self.add_person)
        layout.addWidget(self.add_button)

        # Delete button
        self.delete_button = QPushButton("Delete Person")
        self.delete_button.clicked.connect(self.delete_person)
        layout.addWidget(self.delete_button)

        # List of members
        self.club_people_list = QListWidget()
        layout.addWidget(QLabel("Club Members"))
        layout.addWidget(self.club_people_list)

        self.club_people_tab.setLayout(layout)

    def create_pc_grid(self, num_pcs):
        for i in range(num_pcs):
            pc_button = QPushButton(f"PC {i + 1} (Free)")
            pc_button.setStyleSheet(
                "background-color: lightgreen; font-size: 16px; padding: 10px;"
            )
            pc_button.clicked.connect(lambda _, pc=i: self.toggle_pc_status(pc))
            self.pc_layout.addWidget(pc_button)
            self.pc_statuses[i] = {
                "status": "Free",
                "excel_row": None,
            }  # Initially, all PCs are free

    def toggle_pc_status(self, pc_id):
        if self.pc_statuses[pc_id]["status"] == "Free":
            if not self.club_people_list.count():
                QMessageBox.warning(
                    self,
                    "No Members",
                    "No club members available. Please add members first.",
                )
                return

            # Prompt for club member selection
            dialog = QDialog(self)
            dialog.setWindowTitle("Select Club Member")

            layout = QVBoxLayout()
            self.member_combobox = QComboBox()
            for i in range(self.club_people_list.count()):
                self.member_combobox.addItem(self.club_people_list.item(i).text())

            layout.addWidget(QLabel("Select Club Member:"))
            layout.addWidget(self.member_combobox)

            buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            buttons.accepted.connect(lambda: self.sign_in_member(pc_id, dialog))
            buttons.rejected.connect(dialog.reject)
            layout.addWidget(buttons)

            dialog.setLayout(layout)
            dialog.exec_()

        elif self.pc_statuses[pc_id]["status"] == "In Use":
            sign_out_time = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss")

            pc_button = self.pc_layout.itemAt(pc_id).widget()
            pc_button.setText(f"PC {pc_id + 1} (Free)")
            pc_button.setStyleSheet(
                "background-color: lightgreen; font-size: 16px; padding: 10px;"
            )
            self.pc_statuses[pc_id]["status"] = "Free"

            self.log_sign_out_to_excel(
                self.pc_statuses[pc_id]["excel_row"], sign_out_time
            )

    def sign_in_member(self, pc_id, dialog):
        selected_member = self.member_combobox.currentText()
        student_id = selected_member.split("(")[-1].strip(")")
        student_name = selected_member.split("(")[0].strip()
        sign_in_time = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss")

        pc_button = self.pc_layout.itemAt(pc_id).widget()
        pc_button.setText(f"PC {pc_id + 1} (In Use)")
        pc_button.setStyleSheet(
            "background-color: lightcoral; font-size: 16px; padding: 10px;"
        )
        self.pc_statuses[pc_id]["status"] = "In Use"

        self.people_list.addItem(f"{student_name} (PC {pc_id + 1})")
        row = self.log_sign_in_to_excel(student_name, student_id, sign_in_time)
        self.pc_statuses[pc_id]["excel_row"] = row

        dialog.accept()

    def init_log_excel_file(self):
        current_time = QDateTime.currentDateTime().toString("yyyyMMdd_hhmmss")
        log_file = f"PC_Lounge_Log_{current_time}.xlsx"

        if not os.path.exists(log_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "PC Log"
            ws.append(["Name", "Student ID", "Sign-In Time", "Sign-Out Time"])
            wb.save(log_file)

        return log_file

    def log_sign_in_to_excel(self, name, student_id, sign_in_time):
        wb = openpyxl.load_workbook(self.log_file)
        ws = wb.active
        ws.append([name, student_id, sign_in_time, ""])
        row = ws.max_row
        wb.save(self.log_file)
        wb.close()
        return row

    def log_sign_out_to_excel(self, row, sign_out_time):
        wb = openpyxl.load_workbook(self.log_file)
        ws = wb.active
        ws.cell(row=row, column=4).value = sign_out_time
        wb.save(self.log_file)
        wb.close()

    def init_club_excel_file(self):
        club_file = "Club_Members.xlsx"

        if not os.path.exists(club_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Members"
            ws.append(["Student ID", "Name"])
            wb.save(club_file)

        return club_file

    def load_club_members(self):
        self.club_people_list.clear()
        wb = openpyxl.load_workbook(self.club_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            student_id, name = row
            self.club_people_list.addItem(f"{name} ({student_id})")
        wb.close()

    def add_person(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Person")

        layout = QFormLayout()
        student_id_input = QLineEdit()
        name_input = QLineEdit()
        layout.addRow("Student ID:", student_id_input)
        layout.addRow("Name:", name_input)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(
            lambda: self.save_person(student_id_input.text(), name_input.text(), dialog)
        )
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        dialog.setLayout(layout)
        dialog.exec_()

    def save_person(self, student_id, name, dialog):
        if student_id and name:
            wb = openpyxl.load_workbook(self.club_file)
            ws = wb.active
            ws.append([student_id, name])
            wb.save(self.club_file)
            wb.close()
            self.club_people_list.addItem(f"{name} ({student_id})")
            dialog.accept()
        else:
            QMessageBox.warning(
                self, "Input Error", "Both Student ID and Name are required."
            )

    def delete_person(self):
        selected_item = self.club_people_list.currentItem()
        if selected_item:
            student_id = selected_item.text().split("(")[-1].strip(")")
            self.remove_person_from_file(student_id)
            self.club_people_list.takeItem(self.club_people_list.row(selected_item))

    def remove_person_from_file(self, student_id):
        wb = openpyxl.load_workbook(self.club_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        ws.delete_rows(2, ws.max_row - 1)  # Clear all rows except the header

        for row in rows[1:]:
            if row[0] != student_id:
                ws.append(row)

        wb.save(self.club_file)
        wb.close()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = PCLoungeApp()
    window.show()
    sys.exit(app.exec_())
